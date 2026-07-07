import csv
from pathlib import Path

SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".pdf", ".docx", ".xlsx"}
MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024  # 20 MB


def _extract_txt(path: Path) -> dict:
    text = path.read_text(encoding="utf-8", errors="replace")
    return {"raw_text": text, "char_count": len(text)}


def _extract_csv(path: Path) -> dict:
    with path.open(encoding="utf-8-sig", errors="replace", newline="") as f:
        rows = list(csv.DictReader(f))
    return {"rows": rows, "row_count": len(rows)}


def _extract_pdf(path: Path) -> dict:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    raw_text = "\n\n".join(pages)
    return {"raw_text": raw_text, "char_count": len(raw_text), "page_count": len(pages)}


def _extract_docx(path: Path) -> dict:
    from docx import Document

    doc = Document(str(path))
    sections = []
    current_heading = None
    current_lines = []

    def flush():
        if current_heading is not None or current_lines:
            sections.append({
                "heading": current_heading,
                "text": "\n".join(current_lines).strip(),
            })

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        is_heading = para.style is not None and para.style.name and para.style.name.lower().startswith("heading")
        if is_heading:
            flush()
            current_heading = text
            current_lines = []
        else:
            current_lines.append(text)
    flush()

    raw_text = "\n\n".join(s["text"] for s in sections if s["text"])
    return {"sections": sections, "raw_text": raw_text, "char_count": len(raw_text)}


def _extract_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            sheets[sheet_name] = []
            continue
        rows = []
        for row in rows_iter:
            if all(c is None for c in row):
                continue
            rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
        sheets[sheet_name] = rows
    wb.close()

    total_rows = sum(len(r) for r in sheets.values())
    return {"sheets": sheets, "row_count": total_rows}


_EXTRACTORS = {
    ".txt": _extract_txt,
    ".md": _extract_txt,
    ".csv": _extract_csv,
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
}


def extract_requirements(file_path: str) -> dict:
    path = Path(file_path)

    if not path.is_file():
        return {"error": f"File not found: {file_path}"}

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return {
            "error": f"Unsupported file type '{ext}'. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        }

    size = path.stat().st_size
    if size > MAX_FILE_SIZE_BYTES:
        return {"error": f"File too large ({size} bytes). Max is {MAX_FILE_SIZE_BYTES} bytes."}

    try:
        result = _EXTRACTORS[ext](path)
    except Exception as e:
        return {"error": f"Failed to parse {ext} file: {e}"}

    result["file_name"] = path.name
    result["file_type"] = ext.lstrip(".")
    return result
